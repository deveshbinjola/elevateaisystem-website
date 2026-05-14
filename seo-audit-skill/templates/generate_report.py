"""
SEO Audit Report Generator for Coaches
Generates a professional 7-tab Excel report from audit data.

Usage:
    python generate_report.py --data audit_data.json --output report.xlsx

The audit_data.json should contain:
{
    "company_name": "Example Coaching",
    "website_url": "https://example.com",
    "audit_date": "2026-03-02",
    "health_score": 42,
    "keywords": [...],
    "on_page_issues": [...],
    "content_gaps": [...],
    "technical_checks": [...],
    "competitors": [...],
    "action_plan": {"quick_wins": [...], "strategic": [...]}
}
"""

import json
import sys
import argparse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BRAND_BLUE = "1F4E79"
ACCENT_BLUE = "2E75B6"
WHITE = "FFFFFF"
LIGHT_GRAY = "F2F2F2"
RED = "FF4444"
ORANGE = "FF8C00"
YELLOW = "FFD700"
GREEN = "28A745"

header_font = Font(bold=True, color=WHITE, size=11, name="Arial")
header_fill = PatternFill("solid", fgColor=BRAND_BLUE)
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin_border = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)

SEVERITY_FILLS = {
    "Critical": PatternFill("solid", fgColor="FFE0E0"),
    "High": PatternFill("solid", fgColor="FFF0E0"),
    "Medium": PatternFill("solid", fgColor="FFFFF0"),
    "Low": PatternFill("solid", fgColor="E8F5E9"),
    "Pass": PatternFill("solid", fgColor="E8F5E9"),
    "Fail": PatternFill("solid", fgColor="FFE0E0"),
    "Warning": PatternFill("solid", fgColor="FFF0E0"),
}

def style_header_row(ws, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
    ws.freeze_panes = "A2"

def style_data_rows(ws, num_rows, num_cols, severity_col=None):
    alt_fill = PatternFill("solid", fgColor=LIGHT_GRAY)
    for row in range(2, num_rows + 2):
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = Font(size=10, name="Arial")
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = thin_border
            if severity_col and col == severity_col:
                val = str(cell.value or "")
                if val in SEVERITY_FILLS:
                    cell.fill = SEVERITY_FILLS[val]
                    cell.font = Font(size=10, name="Arial", bold=True)
            elif row % 2 == 0:
                cell.fill = alt_fill

def auto_width(ws, num_cols, max_width=45):
    for col in range(1, num_cols + 1):
        max_len = 0
        letter = get_column_letter(col)
        for row in ws.iter_rows(min_col=col, max_col=col):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[letter].width = min(max_len + 4, max_width)

def create_report(data, output_path):
    wb = Workbook()

    # Tab 1: Executive Summary
    ws = wb.active
    ws.title = "Executive Summary"
    ws.sheet_properties.tabColor = BRAND_BLUE

    ws["A1"] = "SEO Audit Report"
    ws["A1"].font = Font(size=18, bold=True, color=BRAND_BLUE, name="Arial")
    ws["A2"] = f"Website: {data['website_url']}"
    ws["A3"] = f"Company: {data['company_name']}"
    ws["A4"] = f"Audit Date: {data['audit_date']}"
    ws["A6"] = "Overall SEO Health Score"
    ws["A6"].font = Font(size=14, bold=True, color=BRAND_BLUE, name="Arial")
    ws["B6"] = data["health_score"]
    ws["B6"].font = Font(size=24, bold=True, color=BRAND_BLUE, name="Arial")
    ws["C6"] = "/ 100"

    row = 8
    ws.cell(row=row, column=1, value="Top Critical Issues").font = Font(size=12, bold=True, color=RED, name="Arial")
    for i, issue in enumerate(data.get("critical_issues", [])[:5], 1):
        ws.cell(row=row + i, column=1, value=f"{i}. {issue}")
    row += len(data.get("critical_issues", [])[:5]) + 2

    ws.cell(row=row, column=1, value="Quick Wins").font = Font(size=12, bold=True, color=GREEN, name="Arial")
    for i, win in enumerate(data.get("quick_wins_summary", [])[:5], 1):
        ws.cell(row=row + i, column=1, value=f"{i}. {win}")

    ws.column_dimensions["A"].width = 60
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 10

    # Tab 2: Keyword Opportunities
    ws2 = wb.create_sheet("Keyword Opportunities")
    ws2.sheet_properties.tabColor = ACCENT_BLUE
    headers = ["Keyword", "Monthly Volume", "Difficulty", "Current Position", "Opportunity Score", "Recommended Action"]
    ws2.append(headers)
    for kw in data.get("keywords", []):
        ws2.append([kw.get("keyword", ""), kw.get("volume", ""), kw.get("difficulty", ""),
                     kw.get("position", "Not ranking"), kw.get("opportunity", ""), kw.get("action", "")])
    style_header_row(ws2, len(headers))
    style_data_rows(ws2, len(data.get("keywords", [])), len(headers))
    auto_width(ws2, len(headers))

    # Tab 3: On-Page Issues
    ws3 = wb.create_sheet("On-Page Issues")
    ws3.sheet_properties.tabColor = "FF4444"
    headers = ["Page URL", "Issue Type", "Severity", "Current State", "Recommended Fix"]
    ws3.append(headers)
    for issue in data.get("on_page_issues", []):
        ws3.append([issue.get("url", ""), issue.get("type", ""), issue.get("severity", ""),
                     issue.get("current", ""), issue.get("fix", "")])
    style_header_row(ws3, len(headers))
    style_data_rows(ws3, len(data.get("on_page_issues", [])), len(headers), severity_col=3)
    auto_width(ws3, len(headers))

    # Tab 4: Content Gaps
    ws4 = wb.create_sheet("Content Gaps")
    ws4.sheet_properties.tabColor = "FF8C00"
    headers = ["Topic", "Search Volume", "Competitor Ranking", "Content Type", "Priority"]
    ws4.append(headers)
    for gap in data.get("content_gaps", []):
        ws4.append([gap.get("topic", ""), gap.get("volume", ""), gap.get("competitor", ""),
                     gap.get("content_type", ""), gap.get("priority", "")])
    style_header_row(ws4, len(headers))
    style_data_rows(ws4, len(data.get("content_gaps", [])), len(headers), severity_col=5)
    auto_width(ws4, len(headers))

    # Tab 5: Technical SEO
    ws5 = wb.create_sheet("Technical SEO")
    ws5.sheet_properties.tabColor = "28A745"
    headers = ["Check", "Status", "Details", "Fix Required"]
    ws5.append(headers)
    for check in data.get("technical_checks", []):
        ws5.append([check.get("check", ""), check.get("status", ""),
                     check.get("details", ""), check.get("fix", "")])
    style_header_row(ws5, len(headers))
    style_data_rows(ws5, len(data.get("technical_checks", [])), len(headers), severity_col=2)
    auto_width(ws5, len(headers))

    # Tab 6: Competitor Comparison
    ws6 = wb.create_sheet("Competitor Comparison")
    ws6.sheet_properties.tabColor = BRAND_BLUE
    headers = ["Metric"] + [c.get("name", f"Competitor {i+1}") for i, c in enumerate(data.get("competitors", []))]
    if not data.get("competitors"):
        headers = ["Metric", "Your Site", "Competitor 1", "Competitor 2"]
    ws6.append(headers)
    for row_data in data.get("competitor_rows", []):
        ws6.append(row_data)
    style_header_row(ws6, len(headers))
    style_data_rows(ws6, len(data.get("competitor_rows", [])), len(headers))
    auto_width(ws6, len(headers))

    # Tab 7: Action Plan
    ws7 = wb.create_sheet("Action Plan")
    ws7.sheet_properties.tabColor = "28A745"
    headers = ["Action", "Category", "Priority", "Effort", "Impact", "Timeline"]
    ws7.append(headers)
    for action in data.get("action_plan", {}).get("quick_wins", []):
        ws7.append([action.get("action", ""), "Quick Win", action.get("priority", ""),
                     action.get("effort", ""), action.get("impact", ""), action.get("timeline", "")])
    for action in data.get("action_plan", {}).get("strategic", []):
        ws7.append([action.get("action", ""), "Strategic", action.get("priority", ""),
                     action.get("effort", ""), action.get("impact", ""), action.get("timeline", "")])
    total_actions = len(data.get("action_plan", {}).get("quick_wins", [])) + len(data.get("action_plan", {}).get("strategic", []))
    style_header_row(ws7, len(headers))
    style_data_rows(ws7, total_actions, len(headers), severity_col=3)
    auto_width(ws7, len(headers))

    wb.save(output_path)
    print(json.dumps({"status": "success", "output": output_path, "tabs": 7}))

if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--data", required=True)
    parser.add_argument("--output", required=True)
    args = parser.parse_args()
    with open(args.data) as f:
        data = json.load(f)
    create_report(data, args.output)
